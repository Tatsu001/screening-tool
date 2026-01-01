#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
投資管理テンプレート - スクリーニングシート自動更新スクリプト

使い方:
    python update_screening.py 投資管理テンプレート.xlsx

機能:
    1. yfinanceで株価・財務データを取得
    2. スクリーニングシートのみ上書き
    3. ポートフォリオに残っている銘柄は保持（背景色でアラート）
    4. その他のシートは変更なし
"""

import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import time
import os
import glob

# tkinterをインポート（ファイル選択ダイアログ用）
try:
    import tkinter as tk
    from tkinter import filedialog
    HAS_TKINTER = True
except ImportError:
    HAS_TKINTER = False
    print("⚠️  tkinterが利用できません。ファイル選択ダイアログは使用できません。")

# yfinanceのインストール確認
try:
    import yfinance as yf
except ImportError:
    print("yfinanceがインストールされていません。")
    print("以下のコマンドでインストールしてください:")
    print("  pip install yfinance")
    input("\nEnterキーで終了...")
    sys.exit(1)

# 色定義
HEADER_COLOR = "2C3E50"
SUBHEADER_COLOR = "34495E"
INPUT_COLOR = "FFF9E6"
WHITE = "FFFFFF"
SUCCESS_COLOR = "D5F4E6"
WARNING_COLOR = "FCF3CF"
DANGER_COLOR = "FADBD8"
PORTFOLIO_ALERT_COLOR = "FFE5CC"  # ポートフォリオ銘柄アラート色（オレンジ）

def select_excel_file():
    """
    GUIファイル選択ダイアログでExcelファイルを選択
    
    Returns:
        str: 選択されたファイルパス（キャンセル時はNone）
    """
    if not HAS_TKINTER:
        return None
    
    # tkinterのルートウィンドウを作成（非表示）
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    # ファイル選択ダイアログを表示
    filepath = filedialog.askopenfilename(
        title='Excelファイルを選択してください',
        filetypes=[
            ('Excel files', '*.xlsx'),
            ('All files', '*.*')
        ],
        initialdir=os.getcwd()
    )
    
    root.destroy()
    
    return filepath if filepath else None

# グロース市場の銘柄リスト（代表的な銘柄）
GROWTH_MARKET_CODES = {
    '4478', '4755', '4477', '4481', '4486', '4488', '3681', '3696',
    '7047', '7048', '7049', '6070', '6098', '6177', '6178', '6180',
    '4385', '4386', '4431', '4433', '4434', '4435', '4436', '4479',
    '2158', '2326', '2379', '2427', '2428', '3923', '3924', '3928',
    '4368', '4371', '4374', '4375', '4376', '4378', '4382', '4384',
}

# スタンダード市場の銘柄リスト（代表的な銘柄）
STANDARD_MARKET_CODES = {
    '1515', '1518', '1719', '1720', '1721', '1766', '1770', '1780',
    '5401', '5410', '5411', '5444', '5445', '5449', '5451', '5471',
}

def get_market_category(info, ticker_code):
    """
    市場区分を取得
    
    Args:
        info: yfinanceのinfo辞書
        ticker_code: 銘柄コード（4桁）
    
    Returns:
        str: プライム/スタンダード/グロース
    """
    # まず銘柄コードから判定（最も確実）
    if ticker_code in GROWTH_MARKET_CODES:
        return 'グロース'
    
    if ticker_code in STANDARD_MARKET_CODES:
        return 'スタンダード'
    
    # yfinanceのデータから判定を試みる
    exchange = str(info.get('exchange', '')).upper()
    quote_type = str(info.get('quoteType', '')).upper()
    long_name = str(info.get('longName', '')).lower()
    
    # グロース市場のキーワード判定
    if any(keyword in long_name for keyword in ['growth', 'mothers', 'jasdaq growth']):
        return 'グロース'
    
    # 市場情報から判定
    if 'market' in info:
        market_info = str(info.get('market', '')).lower()
        if 'growth' in market_info or 'mothers' in market_info:
            return 'グロース'
        elif 'standard' in market_info:
            return 'スタンダード'
        elif 'prime' in market_info:
            return 'プライム'
    
    # デフォルトはプライム（主要企業は大抵プライム）
    return 'プライム'

def get_stock_data(ticker_code):
    """
    yfinanceで株価・財務データを取得
    
    Args:
        ticker_code: 銘柄コード（例: 7203）
    
    Returns:
        dict: 取得したデータ（データがない場合は'-'）
    """
    try:
        # 日本株は .T を付ける
        ticker = f"{ticker_code}.T"
        stock = yf.Ticker(ticker)
        info = stock.info
        
        # 市場区分を取得（銘柄コードを渡す）
        market = get_market_category(info, ticker_code)
        
        # 自己資本比率を計算
        equity_ratio = None
        if info.get('totalStockholderEquity') and info.get('totalAssets'):
            total_equity = info.get('totalStockholderEquity')
            total_assets = info.get('totalAssets')
            if total_assets > 0:
                equity_ratio = (total_equity / total_assets) * 100
        
        # 売買代金を計算（億円単位）
        trading_value = None
        if info.get('averageVolume') and info.get('currentPrice'):
            avg_volume = info.get('averageVolume')
            current_price = info.get('currentPrice')
            # 出来高 × 株価 ÷ 100,000,000 = 億円
            trading_value = (avg_volume * current_price) / 100000000
        
        # ROE
        roe = info.get('returnOnEquity')
        if roe is not None:
            roe = roe * 100  # パーセント表記
        
        # 売上成長率
        revenue_growth = info.get('revenueGrowth')
        if revenue_growth is not None:
            revenue_growth = revenue_growth * 100  # パーセント表記
        
        # データを辞書形式で返す（Noneの場合は'-'に変換）
        data = {
            'name': info.get('longName', info.get('shortName', '-')),
            'market': market,
            'market_cap': info.get('marketCap'),
            'equity_ratio': equity_ratio,
            'trading_value': trading_value,
            'trailing_pe': info.get('trailingPE'),
            'price_to_book': info.get('priceToBook'),
            'return_on_equity': roe,
            'revenue_growth': revenue_growth,
        }
        
        return data
        
    except Exception as e:
        print(f"  ⚠️  {ticker_code}: データ取得エラー - {str(e)}")
        return None



def format_value(value, format_type='number', decimals=1):
    """
    値をフォーマット（Noneの場合は'-'を返す）
    
    Args:
        value: フォーマットする値
        format_type: 'number', 'percent', 'currency'
        decimals: 小数点以下の桁数
    
    Returns:
        フォーマットされた値または'-'
    """
    if value is None:
        return '-'
    
    try:
        if format_type == 'number':
            return round(value, decimals)
        elif format_type == 'percent':
            return round(value, decimals)
        elif format_type == 'currency':
            return round(value, 0)
        else:
            return value
    except:
        return '-'

def get_screening_stocks(wb):
    """
    スクリーニング銘柄シートから銘柄コードリストを取得
    
    Args:
        wb: openpyxlのワークブック
    
    Returns:
        list: 銘柄コードのリスト
    """
    if 'スクリーニング銘柄' not in wb.sheetnames:
        print("❌ エラー: 'スクリーニング銘柄'シートが見つかりません")
        return []
    
    ws = wb['スクリーニング銘柄']
    stock_codes = []
    
    # A列の2行目以降から銘柄コードを取得
    for row in range(2, 100):  # 最大98銘柄
        code = ws[f'A{row}'].value
        if code and str(code).strip():
            stock_codes.append(str(code).strip())
        elif not code:
            # 空欄が出たら終了
            break
    
    return stock_codes

def get_portfolio_stocks(wb):
    """
    ポートフォリオシートから保有銘柄のコードリストを取得
    
    Args:
        wb: openpyxlのワークブック
    
    Returns:
        set: 保有銘柄コードのセット
    """
    portfolio_stocks = set()
    
    if 'ポートフォリオ' not in wb.sheetnames:
        return portfolio_stocks
    
    ws = wb['ポートフォリオ']
    
    # 7行目から11行目まで（データ行）
    for row in range(7, 12):
        code = ws[f'A{row}'].value
        if code and str(code).strip():
            portfolio_stocks.add(str(code).strip())
    
    return portfolio_stocks

def update_screening_sheet(filepath, stock_codes):
    """
    スクリーニングシートを更新
    
    Args:
        filepath: Excelファイルパス
        stock_codes: 更新する銘柄コードのリスト
    """
    print(f"\n📊 ファイルを読み込み中: {filepath}")
    
    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        print(f"❌ エラー: ファイルが見つかりません - {filepath}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ エラー: ファイルの読み込みに失敗 - {str(e)}")
        sys.exit(1)
    
    if 'スクリーニング' not in wb.sheetnames:
        print("❌ エラー: 'スクリーニング'シートが見つかりません")
        sys.exit(1)
    
    ws = wb['スクリーニング']
    
    # ポートフォリオの保有銘柄を取得
    portfolio_stocks = get_portfolio_stocks(wb)
    print(f"\n🔍 ポートフォリオ保有銘柄: {len(portfolio_stocks)}銘柄")
    if portfolio_stocks:
        print(f"   {', '.join(sorted(portfolio_stocks))}")
    
    # 前回のスクリーニングシートから既存銘柄とL列以降のデータを保存
    print(f"\n📋 既存データを読み込み中...")
    existing_data = {}  # {銘柄コード: {row_data: L列以降のデータ}}
    
    for row in range(6, 21):  # 6～20行目
        code = ws[f'A{row}'].value
        if code and str(code).strip():
            code = str(code).strip()
            # L列以降（12列目以降）のデータを保存
            row_data = {}
            for col in range(12, 25):  # L列(12)～X列(24)
                cell = ws.cell(row=row, column=col)
                row_data[col] = {
                    'value': cell.value,
                    'fill': cell.fill.copy() if cell.fill else None,
                    'font': cell.font.copy() if cell.font else None,
                    'alignment': cell.alignment.copy() if cell.alignment else None,
                    'border': cell.border.copy() if cell.border else None,
                    'number_format': cell.number_format,
                }
            existing_data[code] = row_data
            print(f"   {code}: L列以降のデータを保存")
    
    # データの最終行を見つける（新規行のテンプレート用）
    template_row = None
    for row in range(6, 21):
        code = ws[f'A{row}'].value
        if not code or not str(code).strip():
            template_row = row
            break
    if template_row is None:
        template_row = 21  # 見つからない場合は21行目
    
    print(f"\n📝 テンプレート行: {template_row}行目")
    
    # 統合リストを作成
    stock_codes_set = set(stock_codes)
    portfolio_only = portfolio_stocks - stock_codes_set
    
    unified_list = list(stock_codes) + list(portfolio_only)
    
    print(f"\n📊 統合リスト: {len(unified_list)}銘柄")
    print(f"   スクリーニング銘柄: {len(stock_codes)}銘柄")
    print(f"   ポートフォリオのみ: {len(portfolio_only)}銘柄")
    
    # スタイル定義
    input_fill = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type='solid')
    alert_fill = PatternFill(start_color=PORTFOLIO_ALERT_COLOR, end_color=PORTFOLIO_ALERT_COLOR, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # A～K列のみクリア（L列以降は触らない）
    print(f"\n🧹 A～K列をクリア中...")
    for row in range(6, 21):
        for col in range(1, 12):  # A列(1)～K列(11)
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = openpyxl.styles.PatternFill(fill_type=None)
    
    print(f"\n📡 株価データを取得中...")
    print("=" * 60)
    
    # データ開始行
    current_row = 6
    portfolio_alerts = []
    
    # 統合リストの各銘柄を処理
    for idx, code in enumerate(unified_list, start=1):
        code = str(code).strip()
        
        print(f"\n[{idx}/{len(unified_list)}] {code}")
        
        # ポートフォリオにあるが今回のリストにない = オレンジ色
        is_portfolio_alert = code in portfolio_only
        if is_portfolio_alert:
            print(f"  ⚠️  ポートフォリオ保有中（スクリーニング対象外）")
            portfolio_alerts.append(code)
        
        # yfinanceでデータ取得
        print(f"  取得中...", end=" ")
        data = get_stock_data(code)
        
        if data is None:
            print("スキップ")
            current_row += 1
            continue
        
        print("✓")
        
        # 新規銘柄の場合、テンプレート行から書式・入力規則をコピー
        is_new_stock = code not in existing_data
        if is_new_stock:
            print(f"  📋 新規銘柄: テンプレート行から書式をコピー")
            # L列以降の書式・入力規則をコピー（値はコピーしない）
            for col in range(12, 25):  # L列(12)～X列(24)
                template_cell = ws.cell(row=template_row, column=col)
                target_cell = ws.cell(row=current_row, column=col)
                
                # 値はコピーしない（空欄のまま）
                target_cell.value = None
                
                # 書式をコピー
                if template_cell.fill:
                    target_cell.fill = template_cell.fill.copy()
                if template_cell.font:
                    target_cell.font = template_cell.font.copy()
                if template_cell.alignment:
                    target_cell.alignment = template_cell.alignment.copy()
                if template_cell.border:
                    target_cell.border = template_cell.border.copy()
                if template_cell.number_format:
                    target_cell.number_format = template_cell.number_format
        
        # A～K列を書き込み
        row = current_row
        
        # A列: 銘柄コード
        ws[f'A{row}'] = code
        ws[f'A{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].border = thin_border
        
        # B列: 銘柄名
        name = data['name'] if data['name'] and data['name'] != '-' else '-'
        ws[f'B{row}'] = name
        ws[f'B{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'B{row}'].alignment = center_align
        ws[f'B{row}'].border = thin_border
        
        # C列: 市場区分
        ws[f'C{row}'] = data.get('market', 'プライム')
        ws[f'C{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'C{row}'].alignment = center_align
        ws[f'C{row}'].border = thin_border
        
        # D列: 時価総額
        market_cap = format_value(data['market_cap'] / 100000000 if data['market_cap'] else None, 'currency')
        ws[f'D{row}'] = market_cap
        if market_cap != '-':
            ws[f'D{row}'].number_format = '#,##0'
        ws[f'D{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'D{row}'].alignment = center_align
        ws[f'D{row}'].border = thin_border
        
        # E列: 自己資本比率
        equity_ratio = format_value(data['equity_ratio'], 'number', 1)
        ws[f'E{row}'] = equity_ratio
        if equity_ratio != '-':
            ws[f'E{row}'].number_format = '0.0'
        ws[f'E{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'E{row}'].alignment = center_align
        ws[f'E{row}'].border = thin_border
        
        # F列: 売買代金
        trading_value = format_value(data['trading_value'], 'currency')
        ws[f'F{row}'] = trading_value
        if trading_value != '-':
            ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'F{row}'].alignment = center_align
        ws[f'F{row}'].border = thin_border
        
        # G列: PER
        per = format_value(data['trailing_pe'], 'number', 1)
        ws[f'G{row}'] = per
        if per != '-':
            ws[f'G{row}'].number_format = '0.0'
        ws[f'G{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'G{row}'].alignment = center_align
        ws[f'G{row}'].border = thin_border
        
        # H列: PBR
        pbr = format_value(data['price_to_book'], 'number', 1)
        ws[f'H{row}'] = pbr
        if pbr != '-':
            ws[f'H{row}'].number_format = '0.0'
        ws[f'H{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'H{row}'].alignment = center_align
        ws[f'H{row}'].border = thin_border
        
        # I列: バリュースコア（数式 - 触らない）
        
        # J列: 売上成長率（自動取得）✨
        revenue_growth = format_value(data['revenue_growth'], 'percent', 1)
        ws[f'J{row}'] = revenue_growth
        if revenue_growth != '-':
            ws[f'J{row}'].number_format = '0.0'
        ws[f'J{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'J{row}'].alignment = center_align
        ws[f'J{row}'].border = thin_border
        
        # K列: ROE（自動取得）✨
        roe = format_value(data['return_on_equity'], 'percent', 1)
        ws[f'K{row}'] = roe
        if roe != '-':
            ws[f'K{row}'].number_format = '0.0'
        ws[f'K{row}'].fill = alert_fill if is_portfolio_alert else input_fill
        ws[f'K{row}'].alignment = center_align
        ws[f'K{row}'].border = thin_border
        
        # L列以降: 既存データがあれば復元（数式・手動入力を保持）
        if code in existing_data:
            print(f"  📋 L列以降のデータを復元")
            for col, cell_data in existing_data[code].items():
                cell = ws.cell(row=row, column=col)
                cell.value = cell_data['value']
                if cell_data['fill']:
                    cell.fill = cell_data['fill']
                if cell_data['font']:
                    cell.font = cell_data['font']
                if cell_data['alignment']:
                    cell.alignment = cell_data['alignment']
                if cell_data['border']:
                    cell.border = cell_data['border']
                if cell_data['number_format']:
                    cell.number_format = cell_data['number_format']
        
        current_row += 1
        
        # API制限を避けるため少し待機
        time.sleep(0.5)
    
    # ファイルを保存
    print("\n" + "=" * 60)
    print(f"💾 ファイルを保存中...")
    
    try:
        wb.save(filepath)
        print(f"✅ 保存完了: {filepath}")
    except Exception as e:
        print(f"❌ エラー: ファイルの保存に失敗 - {str(e)}")
        sys.exit(1)
    
    # サマリー表示
    print("\n" + "=" * 60)
    print("📊 更新サマリー")
    print("=" * 60)
    print(f"更新銘柄数: {len(unified_list)}銘柄")
    print(f"  - スクリーニング銘柄: {len(stock_codes)}銘柄")
    print(f"  - ポートフォリオのみ: {len(portfolio_only)}銘柄")
    
    if portfolio_alerts:
        print(f"\n⚠️  ポートフォリオ保有中（スクリーニング対象外）:")
        for code in portfolio_alerts:
            print(f"   - {code}")
        print(f"\n注意: これらの銘柄は売却を検討してください。")
    
    print("\n✅ スクリーニングシート更新完了!")


def main():
    print("\n" + "=" * 60)
    print(f"💾 ファイルを保存中...")
    
    try:
        wb.save(filepath)
        print(f"✅ 保存完了: {filepath}")
    except Exception as e:
        print(f"❌ エラー: ファイルの保存に失敗 - {str(e)}")
        sys.exit(1)
    
    # サマリー表示
    print("\n" + "=" * 60)
    print("📊 更新サマリー")
    print("=" * 60)
    print(f"更新銘柄数: {len(stock_codes)}銘柄")
    
    if portfolio_alerts:
        print(f"\n⚠️  ポートフォリオ保有中の銘柄（オレンジ色背景）:")
        for code in portfolio_alerts:
            print(f"   - {code}")
        print(f"\n注意: これらの銘柄はポートフォリオに残っています。")
        print(f"      売却済みの場合はポートフォリオシートから削除してください。")
    
    print("\n✅ スクリーニングシート更新完了!")

def main():
    """メイン処理"""
    print("=" * 60)
    print("📊 投資管理テンプレート - スクリーニングシート自動更新")
    print("=" * 60)
    
    filepath = None
    
    # コマンドライン引数がある場合
    if len(sys.argv) >= 2:
        filepath = sys.argv[1]
        print(f"\n📁 指定されたファイル: {filepath}")
    else:
        # GUIファイル選択ダイアログを表示
        if HAS_TKINTER:
            print("\n📁 ファイル選択ダイアログを開きます...")
            print("   （ダイアログが表示されない場合は、タスクバーを確認してください）")
            
            filepath = select_excel_file()
            
            if filepath:
                print(f"✅ 選択されたファイル: {filepath}")
            else:
                print("❌ ファイルが選択されませんでした")
        
        # GUIが使えないか、キャンセルされた場合は自動検出
        if not filepath:
            print("\n📁 Excelファイルを自動検出します...")
            
            # 候補となるファイル名
            candidates = [
                'investment_template.xlsx',
                '投資管理テンプレート.xlsx',
                '投資管理テンプレート_配列数式版.xlsx',
            ]
            
            # カレントディレクトリで検索
            for candidate in candidates:
                if os.path.exists(candidate):
                    filepath = candidate
                    print(f"✅ 発見: {filepath}")
                    break
            
            # 見つからない場合、xlsxファイルを全て表示
            if not filepath:
                xlsx_files = glob.glob('*.xlsx')
                if xlsx_files:
                    print("\n以下のExcelファイルが見つかりました:")
                    for i, f in enumerate(xlsx_files, 1):
                        print(f"  {i}. {f}")
                    
                    print("\n使用するファイル番号を入力してください:")
                    try:
                        choice = int(input("番号: ").strip())
                        if 1 <= choice <= len(xlsx_files):
                            filepath = xlsx_files[choice - 1]
                            print(f"✅ 選択: {filepath}")
                        else:
                            print("❌ エラー: 無効な番号です")
                            input("\nEnterキーで終了...")
                            sys.exit(1)
                    except (ValueError, EOFError):
                        print("❌ エラー: 無効な入力です")
                        input("\nEnterキーで終了...")
                        sys.exit(1)
                else:
                    print("\n❌ エラー: Excelファイルが見つかりません")
                    print("\n以下のいずれかのファイルを同じフォルダに配置してください:")
                    print("  - investment_template.xlsx")
                    print("  - 投資管理テンプレート.xlsx")
                    input("\nEnterキーで終了...")
                    sys.exit(1)
    
    # ファイルの存在確認
    if not os.path.exists(filepath):
        print(f"\n❌ エラー: ファイルが見つかりません - {filepath}")
        input("\nEnterキーで終了...")
        sys.exit(1)
    
    # Excelファイルを開いて銘柄リストを取得
    print(f"\n📊 ファイルを読み込み中: {filepath}")
    
    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        print(f"❌ エラー: ファイルの読み込みに失敗 - {str(e)}")
        input("\nEnterキーで終了...")
        sys.exit(1)
    
    # スクリーニング銘柄シートから銘柄コードを取得
    stock_codes = get_screening_stocks(wb)
    wb.close()
    
    if not stock_codes:
        print("\n❌ エラー: 'スクリーニング銘柄'シートに銘柄コードが入力されていません")
        print("\n手順:")
        print("1. Excelファイルを開く")
        print("2. 'スクリーニング銘柄'シートのA列（2行目以降）に銘柄コードを入力")
        print("3. 保存してから再実行")
        input("\nEnterキーで終了...")
        sys.exit(1)
    
    # 重複を削除
    stock_codes = list(dict.fromkeys(stock_codes))
    
    print(f"\n✅ {len(stock_codes)}銘柄を更新します")
    print(f"   {', '.join(stock_codes)}")
    
    # 確認
    try:
        confirm = input("\n続行しますか？ (y/N): ").strip().lower()
    except EOFError:
        confirm = 'n'
    
    if confirm not in ['y', 'yes']:
        print("キャンセルしました")
        input("\nEnterキーで終了...")
        sys.exit(0)
    
    # スクリーニングシートを更新
    try:
        update_screening_sheet(filepath, stock_codes)
    except Exception as e:
        print(f"\n❌ エラーが発生しました: {str(e)}")
        import traceback
        traceback.print_exc()
    
    # 終了前に待機
    input("\nEnterキーで終了...")

if __name__ == "__main__":
    main()
